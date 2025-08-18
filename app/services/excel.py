# backend/app/services/excel.py
import io
import os
from pathlib import Path
from typing import List, Optional, Tuple, Union
from datetime import date, datetime
from openpyxl import load_workbook

# ---------- Data container ----------
class DayHours:
    def __init__(self, work_date: date, hours: float):
        self.work_date = work_date
        self.hours = hours

# ---------- Config ----------
def _resolve_template_path() -> Tuple[Path, bool]:
    env_path = os.getenv("EXCEL_TEMPLATE_PATH", "template/Gudipati_Phani_Babu_Timesheet_Week_Ending_08152025.xlsx")
    p = Path(env_path)
    if not p.is_absolute():
        p = Path(__file__).resolve().parents[1] / env_path
    if not p.exists():
        raise FileNotFoundError(f"Excel template not found at: {p}")
    ext = p.suffix.lower()
    if ext == ".xls":
        raise ValueError("Please use .xlsx or .xlsm template (not .xls).")
    keep_vba = (ext == ".xlsm")
    return p, keep_vba

SHEET_FB = os.getenv("EXCEL_SHEET_NAME", "Timesheet")

# ---------- Helpers ----------
def _coerce_date(val: Union[str, date, None]) -> Optional[date]:
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        s = val.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m-%d-%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except Exception:
                pass
    return None

def _date_to_text(d: Optional[date]) -> Optional[str]:
    if not d:
        return None
    # Always return MM-DD-YYYY
    return d.strftime("%m-%d-%Y")

# ---------- Main ----------
def generate_excel(
    employee_name: str,
    designation: str,
    email_primary: str,
    email_secondary: str,
    client_name: Optional[str] = None,
    week_begin: Optional[date] = None,
    week_end: Optional[date] = None,
    days: Optional[List[DayHours]] = None,
) -> bytes:
    """
    Writes into fixed cell positions.
    Dates -> MM-DD-YYYY (text).
    Hours -> numeric.
    Totals -> computed here.
    """
    template_path, keep_vba = _resolve_template_path()
    wb = load_workbook(
        filename=str(template_path),
        data_only=False,   # ✅ KEEP LOGOS/IMAGES
        keep_vba=keep_vba,
        keep_links=False
    )
    ws = wb[SHEET_FB] if SHEET_FB in wb.sheetnames else wb.active

    # ---- Employee block ----
    ws["G2"].value = (employee_name or "").strip()
    ws["G3"].value = (designation or "").strip()
    ws["G4"].value = (email_primary or "").strip()
    ws["G5"].value = (email_secondary or "").strip()

    # ---- Week Beginning & Ending ----
    if week_begin:
        ws["B9"].value = _date_to_text(_coerce_date(week_begin))
        ws["B9"].number_format = "@"
    if week_end:
        ws["C9"].value = _date_to_text(_coerce_date(week_end))
        ws["C9"].number_format = "@"

    # ---- Daily Dates & Hours ----
    date_cells = ["C11", "D11", "E11", "F11", "G11"]
    hour_cells = ["C12", "D12", "E12", "F12", "G12"]

    total_hours = 0.0
    norm: List[DayHours] = []

    if days:
        for d in days:
            if d is None:
                continue
            wd = _coerce_date(getattr(d, "work_date", None) if hasattr(d, "work_date") else d.get("work_date"))
            hrs_raw = getattr(d, "hours", None) if hasattr(d, "hours") else d.get("hours")
            try:
                hrs = None if hrs_raw in (None, "") else round(float(hrs_raw), 2)
            except Exception:
                hrs = None
            if wd:
                norm.append(DayHours(wd, hrs if hrs is not None else 0.0))

    norm.sort(key=lambda x: x.work_date)
    norm = norm[:5]

    for i, entry in enumerate(norm):
        ws[date_cells[i]].value = _date_to_text(entry.work_date)
        ws[date_cells[i]].number_format = "@"
        ws[hour_cells[i]].value = entry.hours
        ws[hour_cells[i]].number_format = "0.00"
        total_hours += entry.hours

    # ---- Totals ----
    regular_hours = min(total_hours, 40.0)
    overtime_hours = max(total_hours - 40.0, 0.0)

    ws["D9"].value = total_hours
    ws["D9"].number_format = "0.00"
    ws["E9"].value = regular_hours
    ws["E9"].number_format = "0.00"

    if "F9" in ws:
        ws["F9"].value = overtime_hours
        ws["F9"].number_format = "0.00"

    # ---- Save ----
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()
