# backend/app/routes/exports.py

import logging
import os
import io
import re
from datetime import date
from typing import Optional, List, Tuple

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from app.services.excel import generate_excel

router = APIRouter()

# ------------------------
# Models
# ------------------------
class DayHours(BaseModel):
    work_date: date
    hours: float

class ExportRequest(BaseModel):
    employee_name: str
    designation: str
    email_primary: str
    email_secondary: Optional[str] = ""
    week_begin: Optional[date] = None
    week_end: Optional[date] = None
    days: Optional[List[DayHours]] = None
    client_name: Optional[str] = None

# ------------------------
# Helpers
# ------------------------
def _sanitize_name(name: str) -> str:
    s = re.sub(r"\s+", "_", name.strip())
    s = re.sub(r"[^\w]", "", s)    # keep only alphanumeric + underscore
    s = re.sub(r"_+", "_", s)
    return s

def _ext_and_mime_from_template() -> Tuple[str, str]:
    tpl = os.getenv("EXCEL_TEMPLATE_PATH", "template/Gudipati_Phani_Babu_Timesheet_Week_Ending_08152025.xlsx").lower()
    if tpl.endswith(".xlsm"):
        return ".xlsm", "application/vnd.ms-excel.sheet.macroEnabled.12"
    return ".xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

def _build_filename(employee_name: str, week_end: Optional[date]) -> str:
    end = week_end or date.today()
    mmddyyyy = end.strftime("%m%d%Y")
    ext, _ = _ext_and_mime_from_template()
    return f"{_sanitize_name(employee_name)}_week_ending_{mmddyyyy}{ext}"

# ------------------------
# Routes
# ------------------------
@router.post("/weekly/download")
def export_weekly_download(req: ExportRequest):
    """
    Stream the generated Excel file.
    - excel.py writes directly to fixed cells.
    - week_begin/week_end/days[] must be provided.
    - Filename: <name>_week_ending_<MMDDYYYY>.(xlsx|xlsm)
    """
    if not req.week_end:
        req.week_end = date.today()

    if not req.week_begin or not req.week_end or not req.days:
        raise HTTPException(
            status_code=400,
            detail="week_begin, week_end, and days[] are required",
        )

    logging.info("Export request: %s", {k: v for k, v in req.dict().items() if k != "days"})

    try:
        xls_bytes = generate_excel(
            employee_name=req.employee_name,
            designation=req.designation,
            email_primary=req.email_primary,
            email_secondary=req.email_secondary or "",
            client_name=req.client_name,
            week_begin=req.week_begin,
            week_end=req.week_end,
            days=req.days,
        )
    except FileNotFoundError as e:
        raise HTTPException(status_code=500, detail=f"Template not found: {e}")
    except KeyError as e:
        raise HTTPException(status_code=500, detail=f"Worksheet not found (check EXCEL_SHEET_NAME): {e}")
    except ValueError as e:
        raise HTTPException(status_code=500, detail=f"Template format issue: {e}")
    except Exception:
        logging.exception("Unexpected error during Excel generation")
        raise HTTPException(status_code=500, detail="Export failed. See server logs for details.")

    filename = _build_filename(req.employee_name, req.week_end)
    _, content_type = _ext_and_mime_from_template()

    return StreamingResponse(
        io.BytesIO(xls_bytes),
        media_type=content_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@router.get("/debug")
def exports_debug():
    """
    Quick environment + template sanity check.
    """
    from pathlib import Path
    info = {}

    tpl_raw = os.getenv("EXCEL_TEMPLATE_PATH", "template/Gudipati_Phani_Babu_Timesheet_Week_Ending_08152025.xlsx")
    sheet = os.getenv("EXCEL_SHEET_NAME", "Timesheet")

    p = Path(tpl_raw)
    if not p.is_absolute():
        p = Path(__file__).resolve().parents[2] / tpl_raw

    info["EXCEL_TEMPLATE_PATH_raw"] = tpl_raw
    info["resolved_template_path"] = str(p)
    info["template_exists"] = p.exists()
    info["template_suffix"] = p.suffix
    info["EXCEL_SHEET_NAME"] = sheet

    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(p), data_only=True, keep_vba=(p.suffix.lower() == ".xlsm"), keep_links=False)
        info["workbook_sheetnames"] = wb.sheetnames
        info["defined_names"] = "cleared in excel.py"
        info["can_open_workbook"] = True
    except Exception as e:
        info["can_open_workbook"] = False
        info["open_error"] = str(e)

    return info
